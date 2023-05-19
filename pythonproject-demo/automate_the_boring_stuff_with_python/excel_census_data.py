import openpyxl
import pprint


print('Reading File ...')
wb = openpyxl.load_workbook('data/censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')

countryData = {}

print('Reading Rows ...')
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    country = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    if countryData.get(state, None) is None:
        countryData[state] = {country: {'tracts': 0, 'pop': 0}}
    else:
        if countryData[state].get(country, None) is None:
            countryData[state][country] = {'tracts': 1, 'pop': pop}
        else:
            countryData[state][country]['tracts'] += 1
            countryData[state][country]['pop'] += int(pop)

'''
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    country = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countryData.setdefault(state, {})
    countryData[state].setdefault(country, {'tracts': 0, 'pop': 0})
    countryData[state][country]['tracts'] += 1
    countryData[state][country]['pop'] += int(pop)
'''

print('Writing File ...')
resultFile = open('data/census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countryData))
resultFile.close()
print('Done.')


import data.census2010 as census2010

print(census2010.allData['AK']['Anchorage'])
