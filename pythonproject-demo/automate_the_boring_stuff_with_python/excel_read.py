import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook('data/example.xlsx')

print(type(wb))
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(type(sheet))
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)

print('-'*20)

sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print('-'*20)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print('-'*20)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.max_column))
print(get_column_letter(sheet.max_row))

print('-'*20)

print(sheet['A1': 'C3'])
print('')

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print('-'*20)

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

